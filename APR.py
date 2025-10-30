import io
import re
from pathlib import Path
from datetime import datetime

import streamlit as st
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.cell import get_column_letter, column_index_from_string
from openpyxl.utils import get_column_letter


# -------------------------
# CONFIG — CHANGE THESE PATHS
# -------------------------
BASE_DIR = Path(__file__).resolve().parent
TEMPLATE_PATH = BASE_DIR / "Fiche de controle - REFLXP - OF.xlsx"
OUTPUT_DIR    = BASE_DIR / "out"
MAX_PLAN_SLOTS = 6

# Ensure output directory exists
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

# -------------------------
# CORE LOGIC
# -------------------------

def write_into_merged(ws, coord: str, value):
    """
    Write to coord; if it's inside a merged range, redirect to the range's
    top-left (anchor) cell. 
    """
    if value is None:
        return
    for rng in ws.merged_cells.ranges:
        if coord in rng:
            anchor = f"{get_column_letter(rng.min_col)}{rng.min_row}"
            ws[anchor] = value
            return
    ws[coord] = value

def populate_fiche(
    src_path: Path,
    template_path: Path,
    out_path: Path = None,
    row_index: int = 0,
    max_plan: int = MAX_PLAN_SLOTS):
    """
    Populate a COPY of the template with data from the OCR workbook at src_path.
    Returns the path to the saved output file.
    """
    # Load source (general + lines)
    df_gen   = pd.read_excel(src_path, sheet_name="Dossier Technique_general")
    df_lines = pd.read_excel(src_path, sheet_name="Dossier Technique_lines")

    if df_gen.empty:
        raise ValueError("Sheet 'Dossier Technique_general' is empty.")
    row = df_gen.iloc[row_index]

    # Determine plan number from 'Nom du plan'; fallback on OF number or Koncile ID
    num_of = str(row.get("Numéro d' OF") or row.get("Koncile ID") or "")

    # Get the name for download button 
    file_name = f"Fiche de Contrôle - OF n° {num_of}.xlsx"

    # Open template and write to its in-memory copy
    wb = load_workbook(template_path)
    ws = wb["Fiche de contrôle"]  # exact sheet name in your template

    def get_val(col_name: str):
        v = row.get(col_name)
        return None if pd.isna(v) else v

    # Header + green block (safe, unmerged cells on the right)
    assignments = {
        "M1":  get_val("Nom du client"),                    # CLIENT :
        "T1": get_val("Numéro d' OF"),                     # OF N°
        "K2":  get_val("Nom du plan"),                      # N° Plan
        "T2": get_val("Indice plan"),                      # Indice
        "X2": pd.to_datetime(get_val("Date de création")).date()
               if pd.notna(get_val("Date de création")) else None,            # Date

        "X4":  get_val("Matière"),                          # Matière (green)
        "X5":  get_val("Couleur"),                          # Couleur (green)
        "X6":  get_val("Tolérance Générale"),               # Tolérance Générale (green)
        "X7":  get_val("RA mini"),                          # RA Mini (green)
        "X8":  get_val("Cassage Angles Vifs"),              # Cassage Angles Vifs (green)
    }

    for coord, value in assignments.items():
        write_into_merged(ws, coord, value)

    # Grid mappings for up to 6 features
    plan_cells    = ["G9","I9","K9","M9","O9","Q9"]
    plus_cells    = ["G10","I10","K10","M10","O10","Q10"]
    minus_cells   = ["G11","I11","K11","M11","O11","Q11"]
    moyenne_cells = ["G12","I12","K12","M12","O12","Q12"]
    tools_cells   = ["F14","H14","J14","L14","N14","P14"]  # aligned 1..6

    # Normalize column names in df_lines and write
    col_map = {
        "Côtes PLAN": "cote",
        "Tolérance supérieure": "plus",
        "Tolérance inférieure": "minus",
        "Côtes MOYENNES": "moy",
        "Outil de mesure": "outil",
    }
    keep = [c for c in col_map if c in df_lines.columns]
    if keep:
        dfL = df_lines[keep].rename(columns=col_map)
        for i in range(min(max_plan, len(dfL))):
            v_cote  = dfL.iloc[i].get("cote", None)
            v_plus  = dfL.iloc[i].get("plus", None)
            v_minus = dfL.iloc[i].get("minus", None)
            v_moy   = dfL.iloc[i].get("moy", None)
            v_outil = dfL.iloc[i].get("outil", None)

            if pd.notna(v_cote):  ws[plan_cells[i]]    = v_cote
            if pd.notna(v_plus):  ws[plus_cells[i]]    = v_plus
            if pd.notna(v_minus): ws[minus_cells[i]]   = v_minus
            if pd.notna(v_moy):   ws[moyenne_cells[i]] = v_moy
            if pd.notna(v_outil): ws[tools_cells[i]]   = v_outil

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue(), file_name

# -------------------------
# STREAMLIT UI
# -------------------------

st.set_page_config(page_title="Fiche de contrôle - Générateur", page_icon="🧰", layout="centered")
st.title("🧰 Générateur de Fiche de contrôle")
st.caption("Glissez-déposez un fichier Excel (OCR) conforme à la structure *Dossier Technique_2*. "
           "La fiche sera créée et enregistrée et téléchargable.")

uploaded = st.file_uploader(
    "Déposez un ou plusieurs fichiers OCR (.xlsx) avec les feuilles "
    "`Dossier Technique_general` et `Dossier Technique_lines`",
    type=["xlsx"],
    accept_multiple_files=True
)

if uploaded:
    st.divider()
    st.subheader("Résultats")

    for up in uploaded:
        try:
            data = up.read()
            tmp_src = OUTPUT_DIR / f"__tmp_input_{up.name}"
            tmp_src.write_bytes(data)

            # Populate copy of template
            file, file_name = populate_fiche(
                src_path=tmp_src,
                template_path=TEMPLATE_PATH,
                out_path=None,
                row_index=0,
                max_plan=MAX_PLAN_SLOTS
            )

            c1, c2 = st.columns([0.65, 0.35])
            with c1:
                st.success(f"✅ Généré : `{file_name}`")
            with c2:
                st.download_button(
                    label="⬇️ Télécharger",
                    data=file,
                    file_name=file_name,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key=f"dl_{file_name}"
                )

        except Exception as e:
            st.error(f"❌ Erreur sur `{up.name}` : {e}")
        finally:
            # Clean temporary file
            try:
                tmp_src.unlink(missing_ok=True)
            except Exception:
                pass

    st.info(f"Toutes les fiches ont été générées ")
else:
    st.write("Aucun fichier déposé pour l’instant.")
